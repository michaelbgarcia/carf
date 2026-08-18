"""The XLSX control sheet -- ``pipeline/control_sheet.py``.

This file is the review surface and therefore the Part 11 artifact: it is where a
human moves ``review_status`` off ``proposed``. So the tests are about the
properties that make that trustworthy -- every row present, coordinates locked,
suggestions visibly distinguishable from decisions, and read-back that refuses to
drop anything silently.
"""

from __future__ import annotations

import pathlib

import pytest
from openpyxl import load_workbook

from pipeline import control_sheet
from pipeline.control_sheet import (
    COLUMNS,
    LOCKED_COLUMNS,
    SHEET_NAME,
    SUGGESTED_FILL,
    ControlSheetError,
    read_control_sheet,
    to_annotations,
    write_control_sheet,
)
from pipeline.models import (
    AnnotationKind,
    AnnotationSet,
    BBox,
    CRFRow,
    Origin,
    PageGeometry,
    ReviewStatus,
    RowSet,
    SdtmAnnotation,
)

PAGE = PageGeometry(page_index=0, width=612.0, height=792.0, gutter_x=320.0)


def _row(row_id: str, y: float, text_1: str = "", text_2: str = "") -> CRFRow:
    return CRFRow(
        row_id=row_id,
        page_index=0,
        form="Demographics",
        text_1=text_1 or f"question {row_id}",
        text_2=text_2,
        bbox_1=BBox(x0=90.0, y0=y, x1=250.0, y1=y + 12.0),
        bbox_2=BBox(x0=400.0, y0=y, x1=470.0, y1=y + 12.0) if text_2 else None,
    )


@pytest.fixture
def rows() -> RowSet:
    return RowSet(
        source_pdf="synthetic.pdf",
        pages=[PAGE],
        rows=[
            _row("p1_r001", 700.0, "Age (years) at time of consent", "Fixed Unit: years"),
            _row("p1_r002", 660.0, "Sex", "Male"),
            _row("p1_r003", 620.0, "Country of Enrollment"),
        ],
    )


def _annot(row_id: str, slot: int, **kw) -> SdtmAnnotation:
    return SdtmAnnotation(
        annot_id=f"{row_id}_a{slot}",
        row_id=row_id,
        slot=slot,
        page_index=0,
        bbox=BBox(x0=260.0, y0=700.0, x1=330.0, y1=712.0),
        kind=kw.pop("kind", AnnotationKind.VARIABLE),
        **kw,
    )


@pytest.fixture
def proposals(rows) -> AnnotationSet:
    return AnnotationSet(
        source_pdf="synthetic.pdf",
        pages=[PAGE],
        annotations=[
            _annot(
                "p1_r001", 1, domain="DM", variable="AGE",
                source_model="mined precedent (n=7)", suggested=True,
            ),
            _annot(
                "p1_r001", 2, domain="DM", variable="AGEU",
                source_model="mined precedent (n=7)", suggested=True,
            ),
            _annot("p1_r002", 1, domain="DM", variable="SEX", suggested=False),
        ],
    )


def _sheet(path):
    return load_workbook(path)[SHEET_NAME]


def _cells(path) -> dict[str, dict[str, object]]:
    ws = _sheet(path)
    header = [c.value for c in ws[1]]
    out = {}
    for row in ws.iter_rows(min_row=2):
        values = {header[i]: c.value for i, c in enumerate(row)}
        out[values["row_id"]] = values
    return out


# --------------------------------------------------------------------------
# Writing
# --------------------------------------------------------------------------


def test_every_row_gets_a_line_even_with_nothing_proposed(tmp_path, rows, proposals):
    """The reason for extracting text rather than detecting fields.

    A row nothing has mapped shows up as a blank annotation cell for a human to
    fill. A field detector would simply not have produced the row, and its
    absence would be invisible.
    """
    path = write_control_sheet(rows, tmp_path / "cs.xlsx", proposals)
    cells = _cells(path)
    assert set(cells) == {r.row_id for r in rows.rows}
    assert cells["p1_r003"]["anno1"] in (None, "")


def test_columns_are_written_in_the_declared_order(tmp_path, rows):
    path = write_control_sheet(rows, tmp_path / "cs.xlsx")
    assert [c.value for c in _sheet(path)[1]] == COLUMNS


def test_both_annotation_slots_reach_their_own_columns(tmp_path, rows, proposals):
    """AGE in anno1, AGEU in anno2 -- one printed line, two variables."""
    cells = _cells(write_control_sheet(rows, tmp_path / "cs.xlsx", proposals))
    assert cells["p1_r001"]["anno1"] == "AGE"
    assert cells["p1_r001"]["anno2"] == "AGEU"


def test_coordinates_are_written_and_locked(tmp_path, rows, proposals):
    """A hand-edited coordinate silently moves an annotation, with no upside."""
    path = write_control_sheet(rows, tmp_path / "cs.xlsx", proposals)
    ws = _sheet(path)
    header = [c.value for c in ws[1]]

    assert ws.protection.sheet is True
    for row in ws.iter_rows(min_row=2):
        for i, cell in enumerate(row):
            name = header[i]
            if name in LOCKED_COLUMNS:
                assert cell.protection.locked, f"{name} is editable"
            else:
                assert not cell.protection.locked, f"{name} is locked"

    assert _cells(path)["p1_r001"]["coord1"], "coord1 was not written"


def test_suggested_cells_are_greyed_and_decided_ones_are_not(tmp_path, rows, proposals):
    """The signal a reviewer scans for: what still needs looking at.

    Without it, a mapping mined from prior CRFs is indistinguishable from one a
    person chose, and "review the sheet" stops meaning anything.
    """
    path = write_control_sheet(rows, tmp_path / "cs.xlsx", proposals)
    ws = _sheet(path)
    header = [c.value for c in ws[1]]
    anno1 = header.index("anno1") + 1

    by_row = {}
    for row in ws.iter_rows(min_row=2):
        by_row[row[header.index("row_id")].value] = row

    suggested = by_row["p1_r001"][anno1 - 1]
    decided = by_row["p1_r002"][anno1 - 1]
    assert suggested.fill.fgColor.rgb == SUGGESTED_FILL.fgColor.rgb
    assert decided.fill.fgColor.rgb != SUGGESTED_FILL.fgColor.rgb


def test_provenance_is_recorded_per_row(tmp_path, rows, proposals):
    cells = _cells(write_control_sheet(rows, tmp_path / "cs.xlsx", proposals))
    assert cells["p1_r001"]["source"] == "mined precedent (n=7)"


def test_review_status_starts_at_proposed(tmp_path, rows, proposals):
    cells = _cells(write_control_sheet(rows, tmp_path / "cs.xlsx", proposals))
    for values in cells.values():
        assert values["review_status"] == ReviewStatus.PROPOSED.value


def test_enum_columns_get_dropdowns(tmp_path, rows):
    """A typo must not be able to invent a review state."""
    path = write_control_sheet(rows, tmp_path / "cs.xlsx")
    formulas = " ".join(dv.formula1 for dv in _sheet(path).data_validations.dataValidation)
    assert "accepted" in formulas
    assert "NotSubmitted" in formulas


# --------------------------------------------------------------------------
# Reading back
# --------------------------------------------------------------------------


def _edit(path, target: str, **values) -> None:
    """Fill in cells the way a reviewer would.

    The row is located by ``target`` before anything is written, so a test can
    legitimately overwrite the ``row_id`` cell itself -- which is what the
    diverged-sheet and fill-down cases need to simulate.
    """
    wb = load_workbook(path)
    ws = wb[SHEET_NAME]
    header = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2):
        if row[header.index("row_id")].value != target:
            continue
        for name, value in values.items():
            row[header.index(name)].value = value
    wb.save(path)


def test_round_trips_a_reviewed_sheet(tmp_path, rows, proposals):
    path = write_control_sheet(rows, tmp_path / "cs.xlsx", proposals)
    _edit(
        path, "p1_r001",
        domain="DM", anno1="AGE", anno2="AGEU", origin="Collected",
        review_status="accepted", reviewed_by="mgarcia",
    )
    out = to_annotations(path, rows)

    slots = {a.slot: a for a in out.annotations if a.row_id == "p1_r001"}
    assert set(slots) == {1, 2}
    assert slots[1].text == "AGE" and slots[2].text == "AGEU"
    assert slots[1].review_status is ReviewStatus.ACCEPTED
    assert slots[1].reviewed_by == "mgarcia"
    assert slots[1].origin is Origin.COLLECTED


def test_a_group_key_folds_its_rows_into_one_annotation(tmp_path, rows):
    """The repeating-annotation shape: the mapping typed once, the key on every
    row of the block."""
    path = write_control_sheet(rows, tmp_path / "cs.xlsx")
    _edit(path, "p1_r001", group="g1", anno1="DSTERM",
          review_status="accepted", reviewed_by="mgarcia")
    _edit(path, "p1_r002", group="g1", review_status="accepted", reviewed_by="mgarcia")
    _edit(path, "p1_r003", group="g1", review_status="accepted", reviewed_by="mgarcia")

    out = to_annotations(path, rows)

    (annot,) = out.annotations
    assert annot.row_id == "p1_r001"
    assert annot.member_row_ids == ["p1_r001", "p1_r002", "p1_r003"]
    assert annot.display_text() == "DSTERM"
    assert annot.group_id == "g1"


def test_a_group_whose_rows_disagree_names_the_spreadsheet_lines(tmp_path, rows):
    """Row ids are the pipeline's join key; a reviewer is looking at a
    spreadsheet, so the error has to name lines they can go to."""
    path = write_control_sheet(rows, tmp_path / "cs.xlsx")
    _edit(path, "p1_r001", group="g1", anno1="DSTERM")
    _edit(path, "p1_r002", group="g1", anno1="DSDECOD")

    with pytest.raises(ControlSheetError) as exc:
        to_annotations(path, rows)
    assert "spreadsheet row(s) 2, 3" in str(exc.value)


def test_a_grouped_annotation_writes_its_key_to_every_member_row(tmp_path, rows):
    """Five rows carrying the key is what makes the group visible and editable:
    clearing one cell is how a reviewer takes that row out of it. Only the
    anchor carries the text -- repeating it would reintroduce in the sheet the
    duplication the group removed from the page."""
    grouped = AnnotationSet(
        source_pdf="synthetic.pdf",
        pages=[PAGE],
        annotations=[
            _annot(
                "p1_r001", 1, text="DSTERM", group_id="g1",
                member_row_ids=["p1_r001", "p1_r002"], suggested=True,
            )
        ],
    )
    cells = _cells(write_control_sheet(rows, tmp_path / "cs.xlsx", grouped))

    assert cells["p1_r001"]["group"] == "g1"
    assert cells["p1_r002"]["group"] == "g1"
    assert cells["p1_r001"]["anno1"] == "DSTERM"
    assert cells["p1_r002"]["anno1"] in (None, "")
    assert cells["p1_r003"]["group"] in (None, "")


def test_an_assigned_group_key_is_greyed_like_any_other_suggestion(tmp_path, rows):
    grouped = AnnotationSet(
        source_pdf="synthetic.pdf",
        pages=[PAGE],
        annotations=[
            _annot(
                "p1_r001", 1, text="DSTERM", group_id="g_p1_r001",
                member_row_ids=["p1_r001", "p1_r002"], suggested=True,
            )
        ],
    )
    path = write_control_sheet(rows, tmp_path / "cs.xlsx", grouped)
    ws = _sheet(path)
    header = [c.value for c in ws[1]]
    column = header.index("group") + 1
    assert ws.cell(row=2, column=column).fill.fgColor.rgb == SUGGESTED_FILL.fgColor.rgb
    assert ws.cell(row=3, column=column).fill.fgColor.rgb == SUGGESTED_FILL.fgColor.rgb


def test_literal_text_wins_over_the_structured_fields(tmp_path, rows):
    """A reviewer's exact wording is what gets drawn.

    ``parse_mapping_text`` recovers domain/variable off it for querying, but
    rendering must never re-derive the string -- that would silently rewrite what
    a person typed.
    """
    path = write_control_sheet(rows, tmp_path / "cs.xlsx")
    _edit(
        path, "p1_r002",
        anno1="VSORRES when VSTESTCD = SYSBP",
        review_status="accepted", reviewed_by="mgarcia",
    )
    annot = next(a for a in to_annotations(path, rows).annotations if a.row_id == "p1_r002")
    assert annot.display_text() == "VSORRES when VSTESTCD = SYSBP"
    # ...and the structured view was still recovered from it.
    assert annot.variable == "VSORRES"
    assert annot.condition == "VSTESTCD = SYSBP"


def test_geometry_comes_from_the_rows_not_the_sheet(tmp_path, rows):
    path = write_control_sheet(rows, tmp_path / "cs.xlsx")
    _edit(path, "p1_r003", anno1="COUNTRY", review_status="accepted", reviewed_by="m")
    annot = next(a for a in to_annotations(path, rows).annotations if a.row_id == "p1_r003")
    assert annot.bbox == rows.by_id("p1_r003").anchor
    assert annot.page_index == 0


def test_a_note_cell_becomes_a_note_annotation(tmp_path, rows):
    """MSG draws commentary with a dashed border, so the kind has to survive."""
    path = write_control_sheet(rows, tmp_path / "cs.xlsx")
    _edit(
        path, "p1_r003",
        note="See protocol section 7.2 for the derivation.",
        review_status="accepted", reviewed_by="m",
    )
    annot = next(a for a in to_annotations(path, rows).annotations if a.row_id == "p1_r003")
    assert annot.kind is AnnotationKind.NOTE
    assert "protocol section 7.2" in annot.display_text()


def test_blank_rows_produce_no_annotations(tmp_path, rows):
    """An unmapped row is a blank line, not an empty annotation."""
    path = write_control_sheet(rows, tmp_path / "cs.xlsx")
    assert to_annotations(path, rows).annotations == []


def test_an_unknown_row_id_fails_loudly(tmp_path, rows):
    """The sheet and the PDF have diverged; a dropped row would be invisible."""
    path = write_control_sheet(rows, tmp_path / "cs.xlsx")
    _edit(path, "p1_r002", row_id="p9_r999", anno1="SEX")
    with pytest.raises(ControlSheetError, match="not in the extracted rows"):
        to_annotations(path, rows)


def test_a_duplicated_row_and_slot_fails_loudly(tmp_path, rows):
    """A fill-down that repeats a row_id would silently discard one of the two."""
    path = write_control_sheet(rows, tmp_path / "cs.xlsx")
    _edit(path, "p1_r002", anno1="SEX")
    _edit(path, "p1_r003", row_id="p1_r002", anno1="SEX")
    with pytest.raises(ControlSheetError, match="appears twice"):
        to_annotations(path, rows)


def test_a_reviewed_row_with_nobody_named_is_rejected(tmp_path, rows):
    """Part 11: a status transition must name the human who made it.

    Enforced by the model; re-raised here with the spreadsheet row number,
    because "row 3" is actionable to a reviewer and a pydantic traceback is not.
    """
    path = write_control_sheet(rows, tmp_path / "cs.xlsx")
    _edit(path, "p1_r002", anno1="SEX", review_status="accepted", reviewed_by="")
    with pytest.raises(ControlSheetError) as exc:
        to_annotations(path, rows)
    assert "reviewed_by" in str(exc.value)
    assert "p1_r002" in str(exc.value)


def test_a_missing_column_is_reported_rather_than_guessed(tmp_path, rows):
    path = write_control_sheet(rows, tmp_path / "cs.xlsx")
    wb = load_workbook(path)
    ws = wb[SHEET_NAME]
    ws.delete_cols(COLUMNS.index("anno1") + 1)
    wb.save(path)
    with pytest.raises(ControlSheetError, match="missing column"):
        read_control_sheet(path)


def test_a_trailing_blank_line_is_not_an_error(tmp_path, rows):
    """Spreadsheets accumulate empty rows; that is not a defect to raise on."""
    path = write_control_sheet(rows, tmp_path / "cs.xlsx")
    wb = load_workbook(path)
    wb[SHEET_NAME].append([None] * len(COLUMNS))
    wb.save(path)
    assert len(read_control_sheet(path)) == len(rows.rows)


def test_password_is_documented_as_a_guard_rail_not_a_control():
    """Worksheet protection is trivially removable, and the module must say so.

    Asserted because a future reader could otherwise mistake it for a security
    boundary and build an access-control argument on top of it -- which in a
    GxP context would be a claim the code cannot support.
    """
    source = pathlib.Path(control_sheet.__file__).read_text(encoding="utf-8")
    assert "not a security control" in source.lower()
