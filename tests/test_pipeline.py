"""Layout, QC stamping, and the whole loop wired together.

The end-to-end test uses ``standin_response.py``, which is NOT a Copilot
reply -- see that module's warning. It proves the pipeline is connected; it
proves nothing about the parser's tolerance of real chat output.

The synthetic CRF is three pages, which collapses to *one* batch under the
default ceiling -- that collapse is itself the thing the batching design exists
to demonstrate: one round trip covering every page, not one per page.
"""

import pymupdf
import pytest

from pipeline import layout, stamp, xfdf
from pipeline.rows import extract_rows
from pipeline.models import NOT_SUBMITTED_TEXT, AnnotationSet, ReviewStatus
from pipeline.parse_response import ingest_response_file
from pipeline.prompt import batch_pages, write_batches
from pipeline.xfdf_to_pdf import xfdf_to_pdf
from standin_response import build_response

from tests.test_roundtrip import read_annots  # noqa: F401  (shared helper)


@pytest.fixture(scope="module")
def rowset(crfs):
    return extract_rows(crfs["acroform"])


@pytest.fixture(scope="module")
def placed(crfs, rowset, tmp_path_factory):
    """A full annotation set, placed, as ingest_response.py would produce."""
    tmp = tmp_path_factory.mktemp("ingest")
    collected = []
    for pages in batch_pages(rowset):
        path = tmp / f"batch-{pages[0]}.csv"
        path.write_text(build_response(rowset, pages), encoding="utf-8")
        collected.extend(ingest_response_file(path, rowset, pages).annotations)
    annots = AnnotationSet(
        source_pdf=rowset.source_pdf, pages=rowset.pages, annotations=collected
    )
    obstacles = layout.text_obstacles(crfs["acroform"])
    return layout.place_annotations(annots, rowset, obstacles=obstacles)


# --- layout ---------------------------------------------------------------


def test_placement_moves_annotations_off_their_own_row(placed, rowset):
    """Otherwise the annotation prints on top of the question it annotates.

    Only annotations with text are placed. One with none -- a row the stand-in
    reply had no mapping for -- keeps its row's own bbox, because nothing gets
    drawn for it and moving it would imply otherwise.
    """
    anchors = {r.row_id: r.anchor for r in rowset.rows}
    for a in placed.annotations:
        if not a.display_text():
            continue
        anchor = anchors[a.row_id]
        overlap_x = min(a.bbox.x1, anchor.x1) - max(a.bbox.x0, anchor.x0)
        overlap_y = min(a.bbox.y1, anchor.y1) - max(a.bbox.y0, anchor.y0)
        assert overlap_x <= 1.0 or overlap_y <= 1.0, f"{a.annot_id} covers its row"


def test_placed_annotations_do_not_overlap_each_other(placed):
    by_page: dict[int, list] = {}
    for a in placed.annotations:
        by_page.setdefault(a.page_index, []).append(a)
    for annots in by_page.values():
        for i, a in enumerate(annots):
            for b in annots[i + 1 :]:
                ox = min(a.bbox.x1, b.bbox.x1) - max(a.bbox.x0, b.bbox.x0)
                oy = min(a.bbox.y1, b.bbox.y1) - max(a.bbox.y0, b.bbox.y0)
                assert ox <= 1.0 or oy <= 1.0, f"{a.annot_id} overlaps {b.annot_id}"


def test_placed_annotations_stay_on_the_page(placed, rowset):
    pages = {p.page_index: p for p in rowset.pages}
    for a in placed.annotations:
        page = pages[a.page_index]
        assert 0 <= a.bbox.x0 and a.bbox.x1 <= page.width
        assert 0 <= a.bbox.y0 and a.bbox.y1 <= page.height


def test_text_obstacles_keep_annotations_off_printed_captions(crfs, rowset, placed):
    """Without this the position annotations printed over Sitting/Supine/Standing."""
    obstacles = layout.text_obstacles(crfs["acroform"])
    hits = 0
    for a in placed.annotations:
        if not a.display_text():
            continue  # never drawn, so it cannot collide with anything
        for o in obstacles[a.page_index]:
            ox = min(a.bbox.x1, o.x1) - max(a.bbox.x0, o.x0)
            oy = min(a.bbox.y1, o.y1) - max(a.bbox.y0, o.y0)
            if ox > 1.0 and oy > 1.0:
                hits += 1
    assert hits == 0


def test_layout_without_obstacles_still_produces_valid_boxes(placed, rowset):
    bare = layout.place_annotations(placed, rowset)
    assert len(bare.annotations) == len(placed.annotations)


# --- QC stamping ----------------------------------------------------------


def test_qc_preview_is_labelled_as_unreviewed(crfs, placed, tmp_path):
    out = stamp.stamp_annotations(crfs["acroform"], placed, tmp_path / "qc.pdf")
    doc = pymupdf.open(out)
    page = doc[0]
    assert "NOT FOR SUBMISSION" in page.get_text()


def test_qc_preview_keeps_annotations_as_markup(crfs, placed, tmp_path):
    out = stamp.stamp_annotations(crfs["acroform"], placed, tmp_path / "qc.pdf")
    assert len(read_annots(out)) > 0


def test_qc_preview_skips_rejected_annotations(crfs, placed, tmp_path):
    rejected = placed.model_copy(
        update={
            "annotations": [
                a.model_copy(
                    update={
                        "review_status": ReviewStatus.REJECTED,
                        "reviewed_by": "mgarcia",
                    }
                )
                for a in placed.annotations
            ]
        }
    )
    out = stamp.stamp_annotations(crfs["acroform"], rejected, tmp_path / "qc.pdf")
    assert read_annots(out) == []


# --- end to end -----------------------------------------------------------


def test_full_loop_from_pdf_to_annotated_pdf(crfs, rowset, tmp_path):
    """extract -> batch -> (stand-in reply) -> ingest -> XFDF -> annotated PDF."""
    manifest = write_batches(rowset, tmp_path)

    collected = []
    for entry in manifest:
        reply = tmp_path / f"resp-batch{entry['batch']}.csv"
        # Mangled the way a chat UI would mangle it.
        reply.write_text(
            build_response(rowset, entry["pages"], as_markdown_table=True, chatty=True),
            encoding="utf-8",
        )
        collected.extend(
            ingest_response_file(reply, rowset, entry["pages"]).annotations
        )

    annots = AnnotationSet(
        source_pdf=rowset.source_pdf, pages=rowset.pages, annotations=collected
    )
    annots = layout.place_annotations(
        annots, rowset, obstacles=layout.text_obstacles(crfs["acroform"])
    )
    xfdf_path = xfdf.write_xfdf(annots, tmp_path / "blankcrf.xfdf")
    final = xfdf_to_pdf(crfs["acroform"], xfdf_path, tmp_path / "annotated.pdf")

    # Every row is answered, and the AGE / AGEU row answers twice -- the
    # variable2 column expands to a second annotation on the same row.
    assert {a.row_id for a in collected} == {r.row_id for r in rowset.rows}
    assert len(collected) > len(rowset.rows), "variable2 did not expand"

    drawn = sum(len(read_annots(final, p.page_index)) for p in rowset.pages)
    with_text = [a for a in collected if a.display_text()]
    assert drawn == len(with_text)


def test_default_batching_needs_only_one_round_trip_for_the_fixture(rowset, tmp_path):
    """The actual point of the redesign, made explicit as an assertion."""
    manifest = write_batches(rowset, tmp_path)
    assert len(manifest) == 1
    assert manifest[0]["pages"] == [p.page_index for p in rowset.pages]


def test_final_pdf_is_not_flattened(crfs, rowset, placed, tmp_path):
    """FDA review tools expect annotations to stay as searchable PDF markup."""
    xfdf_path = xfdf.write_xfdf(placed, tmp_path / "x.xfdf")
    final = xfdf_to_pdf(crfs["acroform"], xfdf_path, tmp_path / "final.pdf")

    doc = pymupdf.open(final)
    page = doc[1]
    types = {a.type[1] for a in page.annots()}
    assert types == {"FreeText"}
    # Searchable: the text is real text, not a rendered image.
    assert "VSORRES" in page.get_text()


def test_conditional_annotations_reach_the_final_pdf(crfs, placed, tmp_path):
    """The --TESTCD pattern is the point of the whole VS grid."""
    xfdf_path = xfdf.write_xfdf(placed, tmp_path / "x.xfdf")
    final = xfdf_to_pdf(crfs["acroform"], xfdf_path, tmp_path / "final.pdf")
    contents = {c for c, _ in read_annots(final, 1)}
    # Un-prefixed, per MSG: the VS domain is in the page legend and the fill.
    assert "VSORRES when VSTESTCD = SYSBP" in contents
    assert "VSORRES when VSTESTCD = RESP" in contents


def test_unmapped_rows_are_visibly_marked_in_the_final_pdf(crfs, placed, tmp_path):
    """A row that maps to nothing still needs a visible mark.

    A reviewer has to be able to tell "deliberately not submitted" from "nobody
    looked at it", and a blank annotation says neither.
    """
    xfdf_path = xfdf.write_xfdf(placed, tmp_path / "x.xfdf")
    final = xfdf_to_pdf(crfs["acroform"], xfdf_path, tmp_path / "final.pdf")
    contents = {c for c, _ in read_annots(final, 0)}
    assert NOT_SUBMITTED_TEXT in contents


# --- grouped annotations, end to end ---------------------------------------


def test_a_grouped_block_reaches_the_final_pdf_as_one_box(crfs, rowset, tmp_path):
    """The whole claim, drawn: a block of rows sharing one mapping produces one
    annotation on the page, not one per row."""
    from pipeline import control_sheet, grouping

    sheet = control_sheet.write_control_sheet(rowset, tmp_path / "cs.xlsx")
    block = [r for r in rowset.rows if r.page_index == 0][10:14]
    _fill_group(sheet, [r.row_id for r in block], "ETHNIC")

    annotations = control_sheet.to_annotations(sheet, rowset)
    placed = layout.place_annotations(annotations, rowset)
    assert grouping.summarize(placed) == (1, len(block))

    final = xfdf_to_pdf(
        crfs["acroform"], xfdf.write_xfdf(placed, tmp_path / "x.xfdf"),
        tmp_path / "final.pdf",
    )
    assert [c for c, _ in read_annots(final, 0)] == ["ETHNIC"]


def test_a_grouped_block_is_not_re_asked_of_copilot(crfs, rowset, tmp_path):
    """Coverage is what the batching reads. Scanning row_id alone would send
    every member but the anchor off to re-derive a mapping already on the
    sheet."""
    from pipeline import control_sheet
    from pipeline.prompt import rows_needing_annotation

    sheet = control_sheet.write_control_sheet(rowset, tmp_path / "cs.xlsx")
    block = [r for r in rowset.rows if r.page_index == 0][10:14]
    _fill_group(sheet, [r.row_id for r in block], "ETHNIC")

    remaining = rows_needing_annotation(
        rowset, control_sheet.to_annotations(sheet, rowset)
    )
    assert not {r.row_id for r in remaining.rows} & {r.row_id for r in block}


def _fill_group(sheet_path, row_ids: list[str], text: str) -> None:
    """Type one mapping and a shared group key, the way a reviewer would."""
    from openpyxl import load_workbook

    from pipeline.control_sheet import SHEET_NAME

    wb = load_workbook(sheet_path)
    ws = wb[SHEET_NAME]
    header = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2):
        if row[header.index("row_id")].value not in row_ids:
            continue
        row[header.index("group")].value = "g1"
        row[header.index("review_status")].value = "accepted"
        row[header.index("reviewed_by")].value = "mgarcia"
        if row[header.index("row_id")].value == row_ids[0]:
            row[header.index("anno1")].value = text
    wb.save(sheet_path)
