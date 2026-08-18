"""The XLSX control sheet -- where a human reviews and completes annotations.

One row per extracted CRF row. Coordinates and text come from extraction and are
locked; ``domain``, ``anno1``, ``anno2``, ``note`` and ``origin`` are the
reviewer's. Cells filled in from mined precedent are shaded grey, so "suggested,
still needs a look" is visible at a glance rather than indistinguishable from
"a person decided this".

Why a spreadsheet is the review surface
---------------------------------------
It replaces two things that were worse. Reviewing XFDF in Acrobat meant one
person at a time, scrolling a several-hundred-page PDF to find every instance of
an annotation that needed changing. And routing every mapping through a Copilot
chat round trip put a text-mangling hop on the critical path for data that is
mostly already known -- the sheet arrives pre-populated, and only the gaps need
a model at all.

Part 11 consequence: **this file is now the artifact a human touches**, so it is
where ``review_status`` moves off ``proposed`` and where ``reviewed_by`` is
recorded. That argument used to attach to the XFDF; it transfers here whole.

anno1 / anno2 are free text, on purpose
---------------------------------------
An annotator wants to type ``VSORRES when VSTESTCD = SYSBP`` into one cell, not
decompose it across ``domain`` / ``variable`` / ``condition`` columns. So the
cells are literal text and that literal text is what gets drawn.
``parse_mapping_text`` recovers the structured fields off it best-effort for
querying and corpus mining, but never feeds them back into rendering -- see
``SdtmAnnotation``'s docstring on which one wins.

Reading back fails loudly
-------------------------
An unknown ``row_id``, a duplicate ``(row_id, slot)``, or a ``review_status``
past ``proposed`` with nobody named all raise rather than being dropped. A
silently discarded annotation is the one defect a reviewer cannot see, and the
sheet is exactly where a stray copy/paste or an autofilled drag produces one --
the same posture as the cell-count check on the markdown-table reply path.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from pipeline.models import (
    AnnotationKind,
    AnnotationSet,
    BBox,
    CRFRow,
    Origin,
    ReviewStatus,
    RowSet,
    SdtmAnnotation,
)

SHEET_NAME = "annotations"

#: Column order. ``row_id`` first because it is the join key a reviewer needs to
#: quote when asking about a row.
COLUMNS = [
    "row_id",
    "page",
    "form",
    "text_1",
    "text_2",
    "coord1",
    "coord2",
    "domain",
    "anno1",
    "anno2",
    "note",
    "origin",
    "source",
    "review_status",
    "reviewed_by",
]

#: Tool output. Locked, because a hand-edited coordinate has no upside and would
#: silently move an annotation somewhere nobody chose.
LOCKED_COLUMNS = {"row_id", "page", "form", "text_1", "text_2", "coord1", "coord2", "source"}

EDITABLE_COLUMNS = [c for c in COLUMNS if c not in LOCKED_COLUMNS]

HEADER_FILL = PatternFill("solid", fgColor="FF1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFFFF")
#: Grey = pre-populated, not yet confirmed by a person.
SUGGESTED_FILL = PatternFill("solid", fgColor="FFD9D9D9")
LOCKED_FILL = PatternFill("solid", fgColor="FFF2F2F2")

COLUMN_WIDTHS = {
    "row_id": 12, "page": 6, "form": 16, "text_1": 52, "text_2": 22,
    "coord1": 26, "coord2": 26, "domain": 8, "anno1": 30, "anno2": 30,
    "note": 40, "origin": 13, "source": 20, "review_status": 14, "reviewed_by": 16,
}

#: Sheet protection password. Not a security control -- worksheet protection is
#: trivially removable and openpyxl itself can strip it. It is a guard rail that
#: makes editing a coordinate take a deliberate act instead of a stray click.
PROTECTION_PASSWORD = "carf"


class ControlSheetError(ValueError):
    """The sheet does not describe the rows it claims to."""


def _fmt_bbox(bbox: Optional[BBox]) -> str:
    if bbox is None:
        return ""
    return ", ".join(f"{v:.2f}" for v in bbox.as_tuple())


def _parse_bbox(value: object) -> Optional[BBox]:
    if value is None or not str(value).strip():
        return None
    parts = [p for p in str(value).replace(",", " ").split() if p]
    if len(parts) != 4:
        raise ControlSheetError(f"expected 4 numbers in a coordinate cell, got {value!r}")
    x0, y0, x1, y1 = (float(p) for p in parts)
    return BBox(x0=min(x0, x1), y0=min(y0, y1), x1=max(x0, x1), y1=max(y0, y1))


# --------------------------------------------------------------------------
# Writing
# --------------------------------------------------------------------------


def _annotations_by_row(
    annotations: Optional[AnnotationSet],
) -> dict[tuple[str, int], SdtmAnnotation]:
    if annotations is None:
        return {}
    out: dict[tuple[str, int], SdtmAnnotation] = {}
    for a in annotations.annotations:
        if a.row_id:
            out[(a.row_id, a.slot)] = a
    return out


def _row_values(
    row: CRFRow, first: Optional[SdtmAnnotation], second: Optional[SdtmAnnotation]
) -> dict[str, object]:
    lead = first or second
    return {
        "row_id": row.row_id,
        "page": row.display_page,
        "form": row.form,
        "text_1": row.text_1,
        "text_2": row.text_2,
        "coord1": _fmt_bbox(row.bbox_1),
        "coord2": _fmt_bbox(row.bbox_2),
        "domain": (lead.domain or "") if lead else "",
        "anno1": first.display_text() if first else "",
        "anno2": second.display_text() if second else "",
        "note": (lead.rationale or "") if lead and lead.kind is AnnotationKind.NOTE else "",
        "origin": (lead.origin.value if lead and lead.origin else ""),
        "source": (lead.source_model if lead else ""),
        "review_status": (lead.review_status.value if lead else ReviewStatus.PROPOSED.value),
        "reviewed_by": (lead.reviewed_by or "") if lead else "",
    }


def write_control_sheet(
    rows: RowSet,
    out_path: str | Path,
    annotations: Optional[AnnotationSet] = None,
) -> Path:
    """Write one sheet row per CRF row, pre-populated from ``annotations``.

    Every extracted row gets a line, including rows nothing has been proposed
    for. That is the point of extracting text rather than detecting fields: an
    unmapped row is visible as a blank annotation cell, where a field detector
    would simply not have produced the row at all.
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    by_row = _annotations_by_row(annotations)
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    ws.append(COLUMNS)
    for i, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=i)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = COLUMN_WIDTHS.get(name, 14)

    for r, row in enumerate(rows.rows, start=2):
        first = by_row.get((row.row_id, 1))
        second = by_row.get((row.row_id, 2))
        values = _row_values(row, first, second)
        for c, name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=r, column=c, value=values[name])
            cell.alignment = Alignment(vertical="top", wrap_text=name in ("text_1", "note"))
            if name in LOCKED_COLUMNS:
                cell.fill = LOCKED_FILL
            else:
                # Constructed rather than copied: openpyxl deprecated
                # StyleProxy.copy(), and a Protection object carries only these
                # two flags, so there is nothing to preserve from the default.
                cell.protection = Protection(locked=False, hidden=False)
            # Grey out what was pre-populated rather than authored, so a reviewer
            # can see what still needs looking at.
            owner = first if name in ("anno1",) else second if name == "anno2" else first or second
            if name in ("domain", "anno1", "anno2", "origin") and owner is not None:
                if owner.suggested and values[name]:
                    cell.fill = SUGGESTED_FILL

    ws.freeze_panes = "A2"
    _add_validation(ws, len(rows.rows))
    ws.protection.password = PROTECTION_PASSWORD
    ws.protection.sheet = True
    ws.protection.enable()

    wb.save(out_path)
    return out_path


def _add_validation(ws, n_rows: int) -> None:
    """Dropdowns for the two enum columns, so a typo cannot invent a state."""
    if n_rows <= 0:
        return
    last = n_rows + 1
    for name, options in (
        ("review_status", [s.value for s in ReviewStatus]),
        ("origin", [""] + [o.value for o in Origin]),
    ):
        letter = get_column_letter(COLUMNS.index(name) + 1)
        dv = DataValidation(
            type="list",
            formula1='"' + ",".join(options) + '"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle=f"invalid {name}",
            error=f"{name} must be one of: " + ", ".join(o for o in options if o),
        )
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}{last}")


# --------------------------------------------------------------------------
# Reading
# --------------------------------------------------------------------------


@dataclass
class SheetRow:
    """One line of the sheet, as read back."""

    row_id: str
    domain: str = ""
    anno1: str = ""
    anno2: str = ""
    note: str = ""
    origin: str = ""
    source: str = ""
    review_status: str = ReviewStatus.PROPOSED.value
    reviewed_by: str = ""
    excel_row: int = 0

    def annotation_texts(self) -> list[tuple[int, str]]:
        """``(slot, text)`` for the filled-in annotation cells."""
        return [(slot, t) for slot, t in ((1, self.anno1), (2, self.anno2)) if t.strip()]


def read_control_sheet(path: str | Path) -> list[SheetRow]:
    """Read the sheet back, without validating it against a ``RowSet`` yet."""
    wb = load_workbook(path, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    missing = [c for c in COLUMNS if c not in header]
    if missing:
        raise ControlSheetError(
            f"{path}: control sheet is missing column(s) {', '.join(missing)}; "
            "it was not produced by write_control_sheet, or columns were deleted"
        )
    index = {name: header.index(name) for name in COLUMNS}

    def text(cells, name: str) -> str:
        value = cells[index[name]].value
        return "" if value is None else str(value).strip()

    out: list[SheetRow] = []
    for excel_row, cells in enumerate(ws.iter_rows(min_row=2), start=2):
        row_id = text(cells, "row_id")
        if not row_id:
            continue  # a trailing blank line is not an error
        out.append(
            SheetRow(
                row_id=row_id,
                domain=text(cells, "domain"),
                anno1=text(cells, "anno1"),
                anno2=text(cells, "anno2"),
                note=text(cells, "note"),
                origin=text(cells, "origin"),
                source=text(cells, "source"),
                review_status=text(cells, "review_status") or ReviewStatus.PROPOSED.value,
                reviewed_by=text(cells, "reviewed_by"),
                excel_row=excel_row,
            )
        )
    return out


def _validate(sheet_rows: Iterable[SheetRow], rows: RowSet) -> None:
    seen: set[tuple[str, int]] = set()
    for sr in sheet_rows:
        if rows.by_id(sr.row_id) is None:
            raise ControlSheetError(
                f"row {sr.excel_row}: row_id {sr.row_id!r} is not in the extracted rows. "
                "The sheet and the PDF have diverged -- re-extract, or check whether a "
                "row_id cell was overwritten by a fill-down"
            )
        for slot, _text in sr.annotation_texts():
            key = (sr.row_id, slot)
            if key in seen:
                raise ControlSheetError(
                    f"row {sr.excel_row}: {sr.row_id!r} slot {slot} appears twice; "
                    "one of the two would be silently discarded"
                )
            seen.add(key)


def to_annotations(
    path: str | Path,
    rows: RowSet,
    source_pdf: Optional[str] = None,
) -> AnnotationSet:
    """Read a reviewed control sheet into annotations positioned on their rows.

    Bboxes come out as the row's own anchor; ``pipeline.layout.place_annotations``
    moves them into the gutter afterwards. Doing that here would bake placement
    into the parse step and make the sheet unreadable without a page geometry.

    A ``review_status`` other than ``proposed`` requires ``reviewed_by``. That is
    enforced by ``SdtmAnnotation`` itself, but the error is re-raised here with
    the offending spreadsheet row number, because "row 214" is actionable to a
    reviewer and a pydantic traceback is not.
    """
    from pipeline.parse_annotated_pdf import parse_mapping_text

    sheet_rows = read_control_sheet(path)
    _validate(sheet_rows, rows)

    out: list[SdtmAnnotation] = []
    for sr in sheet_rows:
        row = rows.by_id(sr.row_id)
        assert row is not None  # _validate

        entries = sr.annotation_texts()
        if sr.note.strip():
            entries.append((max((s for s, _ in entries), default=0) + 1, sr.note.strip()))

        for slot, text in entries:
            is_note = slot > 2 or (not sr.anno1 and not sr.anno2)
            domain, variable, condition, fixed_value = parse_mapping_text(text)
            try:
                out.append(
                    SdtmAnnotation(
                        annot_id=f"{sr.row_id}_a{min(slot, 2)}",
                        row_id=sr.row_id,
                        slot=min(slot, 2),
                        page_index=row.page_index,
                        bbox=row.anchor,
                        kind=AnnotationKind.NOTE if is_note else AnnotationKind.VARIABLE,
                        text=text,
                        domain=domain or sr.domain or None,
                        variable=variable,
                        condition=condition,
                        fixed_value=fixed_value,
                        origin=sr.origin or None,
                        source_model=sr.source or "human, control sheet",
                        suggested=False,
                        review_status=ReviewStatus(sr.review_status),
                        reviewed_by=sr.reviewed_by or None,
                    )
                )
            except ValueError as exc:
                raise ControlSheetError(f"row {sr.excel_row} ({sr.row_id}): {exc}") from exc

    return AnnotationSet(
        source_pdf=source_pdf or rows.source_pdf, pages=rows.pages, annotations=out
    )


__all__ = [
    "COLUMNS",
    "EDITABLE_COLUMNS",
    "LOCKED_COLUMNS",
    "SHEET_NAME",
    "SUGGESTED_FILL",
    "ControlSheetError",
    "SheetRow",
    "read_control_sheet",
    "to_annotations",
    "write_control_sheet",
]
