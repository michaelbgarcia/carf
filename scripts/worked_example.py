#!/usr/bin/env python3
"""Fill in the synthetic CRF's control sheet by hand, so the whole path can run.

NOT PIPELINE OUTPUT, and not a proposal from anything. Every mapping in
``MAPPINGS`` below is **invented by a human** to annotate an invented CRF. Nothing
here may feed a real submission artifact.

It replaces an earlier "illustrative target" script that hand-drew the goal at
hand-picked coordinates, page by page. That existed because placement was a
collision search whose output was hard to predict, so the picture of the goal had
to be drawn separately from the code that would eventually produce it. Placement
is now arithmetic off each row's own anchor, so the goal is just what the pipeline
produces -- and this script exercises the real path instead of imitating it.

Two things it is good for:

* **End to end without a Copilot round trip.** Rows -> control sheet -> reviewed
  sheet -> annotated PDF, all of it real code, on a CRF that ships with the repo.
* **Seeing the MSG styling.** The Demographics form deliberately carries three
  domains (DM, RP, DS) so the per-form encounter-order colouring and the page-top
  legend both have something to do.
* **Seeing grouped annotations.** ``GROUPS`` below puts the Ethnicity and
  Position option blocks under one annotation each, and the instruction
  paragraph under one ``[NOT SUBMITTED]`` -- three shapes of the repeating case,
  drawn once instead of eleven times.

Usage:
    python scripts/worked_example.py [--build-dir build] [--pdf <blank_crf.pdf>]
"""

from __future__ import annotations

import argparse
import subprocess
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook  # noqa: E402

from pipeline import control_sheet, rows as rows_mod  # noqa: E402

DEFAULT_PDF = Path("fixtures/SYNTHETIC_sample_crf_twocol_flat.pdf")
REVIEWER = "worked_example.py (INVENTED - not a real reviewer)"

#: ``text_1 or text_2`` -> ``(domain, anno1, anno2, note, origin)``.
#:
#: Keyed on the row's printed text rather than its ``row_id`` so it survives a
#: layout change to the synthetic CRF: a shifted row keeps its text but gets a new
#: id. Every value is invented.
MAPPINGS: dict[str, tuple[str, str, str, str, str]] = {
    # --- Demographics -----------------------------------------------------
    "Site Identifier": ("DM", "SITEID", "", "", "Collected"),
    "Was this participant a prior screen failure?": (
        "DM", "RESCREEN in SUPPDM", "", "", "Collected",
    ),
    "If Yes, please provide the original participant number (xxxxx-xxxx)": (
        "DM", "PSUBJID in SUPPDM", "", "", "Collected",
    ),
    "Year of Birth (yyyy)": ("DM", "BRTHDTC", "", "", "Collected"),
    # The two-variables-on-one-line case: AGE and its unit, side by side.
    "Age (years) at time of consent": ("DM", "AGE", "AGEU", "", "Collected"),
    "Sex": ("DM", "SEX", "", "", "Collected"),
    # A second domain on the Demographics form, and the findings-condition
    # pattern -- long enough that anno2 has to wrap onto the line below.
    "Is the participant of childbearing potential?": (
        "RP", "RPORRES", "RPTESTCD = CHILDPOT", "", "Collected",
    ),
    "Ethnicity": ("DM", "ETHNIC", "", "", "Collected"),
    "Race (check all that apply)": ("DM", "RACE", "", "", "Collected"),
    # The full-width note that spans the gutter, annotated as a note.
    "When multiple values are selected then RACE = MULTIPLE and individual "
    "responses are RACE1, RACE2, RACEn in SUPPDM": (
        "DM", "", "", "Multiple selections map to RACE1..RACEn in SUPPDM.", "",
    ),
    "American Indian or Alaska Native": (
        "DM", "RACE = AMERICAN INDIAN OR ALASKA NATIVE", "", "", "Collected",
    ),
    "Asian": ("DM", "RACE = ASIAN", "", "", "Collected"),
    "Black or African American": ("DM", "RACE = BLACK OR AFRICAN AMERICAN", "", "", "Collected"),
    "White": ("DM", "RACE = WHITE", "", "", "Collected"),
    "Country of Enrollment": ("DM", "COUNTRY", "", "", "Collected"),
    # A third domain, so the legend has three entries on one form.
    "DEMOGRAPHICS": ("DS", "DSCAT = PROTOCOL MILESTONE", "", "", "Assigned"),
    # --- Vital Signs ------------------------------------------------------
    "Visit Date": ("VS", "VSDTC", "", "", "Collected"),
    "Systolic Blood Pressure": ("VS", "VSORRES", "VSTESTCD = SYSBP", "", "Collected"),
    "Diastolic Blood Pressure": ("VS", "VSORRES", "VSTESTCD = DIABP", "", "Collected"),
    "Pulse Rate": ("VS", "VSORRES", "VSTESTCD = PULSE", "", "Collected"),
    "Body Temperature": ("VS", "VSORRES", "VSTESTCD = TEMP", "", "Collected"),
    "Respiratory Rate": ("VS", "VSORRES", "VSTESTCD = RESP", "", "Collected"),
    "Position": ("VS", "VSPOS", "", "", "Collected"),
}

#: ``text_1 or text_2`` -> group key, for the rows that share one annotation.
#:
#: The repeating case, in the three shapes it actually turns up in:
#:
#: * ``eth`` -- a question row plus its option-only continuation rows. The
#:   mapping is typed once, against "Ethnicity"; the four options carry the key
#:   and an empty cell, and one ``ETHNIC`` box covers the block.
#: * ``pos`` -- the same shape on another form, so the grouped box is drawn
#:   against a different domain's colour.
#: * ``instr`` -- a paragraph of instruction text, seven consecutive rows none
#:   of which is submitted. Without the group this is seven identical grey boxes
#:   down the right of the page.
GROUPS: dict[str, str] = {
    "Ethnicity": "eth",
    "Not Hispanic or Latino": "eth",
    "Not Reported": "eth",
    "Unknown": "eth",
    "Position": "pos",
    "Supine": "pos",
    "Standing": "pos",
    "Complete every page": "instr",
    "any kind.": "instr",
    "line, enter": "instr",
    "original entry": "instr",
    "Where a measurement": "instr",
    "than leaving": "instr",
    "during data review": "instr",
}

#: Rows that carry no submitted data. Matched as a prefix, since page furniture
#: varies its tail ("Page 1 of 3", "Page 2 of 3", ...).
NOT_SUBMITTED_PREFIXES = (
    "SYNTHETIC TEST DATA",
    "Protocol SYNTH-001",
    "Form:",
    "Form Demographics",
    "Form Vital Signs",
    "Form Instructions",
    "Page ",
    "Investigator Initials",
    "Assessor Initials",
    "VITAL SIGNS",
    "INSTRUCTIONS FOR COMPLETION",
    "Complete every page",
    "any kind.",
    "line, enter",
    "original entry",
    "Where a measurement",
    "than leaving",
    "during data review",
)


def _group_of(key: str) -> str:
    """The group key for a row, by exact text then by prefix.

    Prefix matching for the same reason ``NOT_SUBMITTED_PREFIXES`` needs it: the
    instruction paragraph's rows are wrapped lines whose tails are whatever the
    wrap produced.
    """
    if key in GROUPS:
        return GROUPS[key]
    for prefix, group in GROUPS.items():
        if key.startswith(prefix):
            return group
    return ""


def fill_sheet(sheet_path: Path, rows) -> tuple[int, int, int]:
    """Write the invented mappings into an existing control sheet in place.

    Edits the workbook rather than building one, so what gets rendered went
    through exactly the same reader a human-edited sheet does.
    """
    wb = load_workbook(sheet_path)
    ws = wb[control_sheet.SHEET_NAME]
    header = [c.value for c in ws[1]]
    col = {name: header.index(name) + 1 for name in control_sheet.COLUMNS}

    mapped = 0
    not_submitted = 0
    grouped = 0
    for excel_row in range(2, ws.max_row + 1):
        row_id = ws.cell(row=excel_row, column=col["row_id"]).value
        row = rows.by_id(str(row_id)) if row_id else None
        if row is None:
            continue
        key = row.text_1 or row.text_2

        def put(name: str, value: str) -> None:
            ws.cell(row=excel_row, column=col[name], value=value)

        # Written before the mapping branches, and independently of them: most
        # members of a group carry the key and nothing else, so a group cell
        # that only got written for annotated rows would leave four fifths of
        # every block out of its own group.
        group = _group_of(key)
        if group:
            put("group", group)
            grouped += 1

        if key in MAPPINGS:
            domain, anno1, anno2, note, origin = MAPPINGS[key]
            put("domain", domain)
            put("anno1", anno1)
            put("anno2", anno2)
            put("note", note)
            put("origin", origin)
            mapped += 1
        elif key.startswith(NOT_SUBMITTED_PREFIXES):
            put("domain", "")
            put("anno1", "[NOT SUBMITTED]")
            put("origin", "NotSubmitted")
            not_submitted += 1
        elif not group:
            continue

        put("source", "hand-authored (INVENTED example data)")
        put("review_status", "accepted")
        put("reviewed_by", REVIEWER)

    # Protection is re-applied: the sheet this produces should behave like any
    # other reviewed sheet, including its locked columns.
    ws.protection.password = control_sheet.PROTECTION_PASSWORD
    ws.protection.sheet = True
    ws.protection.enable()
    wb.save(sheet_path)
    return mapped, not_submitted, grouped


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--pdf", type=Path, default=DEFAULT_PDF)
    parser.add_argument("--build-dir", type=Path, default=Path("build"))
    args = parser.parse_args(argv)

    if not args.pdf.exists():
        parser.error(f"{args.pdf} not found -- run scripts/make_sample_crf.py first")

    here = Path(__file__).resolve().parent
    print("=== build_sheet.py ===")
    subprocess.run(
        [sys.executable, str(here / "build_sheet.py"), str(args.pdf),
         "--build-dir", str(args.build_dir)],
        check=True,
    )

    sheet = args.build_dir / "control_sheet.xlsx"
    rows = rows_mod.extract_rows(args.pdf)
    mapped, not_submitted, grouped = fill_sheet(sheet, rows)
    print(
        f"\n=== filled {sheet} by hand ===\n"
        f"{mapped} mapped rows, {not_submitted} marked [NOT SUBMITTED], "
        f"{len(rows.rows) - mapped - not_submitted} left blank; "
        f"{grouped} rows share a group key"
    )

    out = args.build_dir / "WORKED_EXAMPLE_annotated_crf.pdf"
    print("\n=== annotate.py ===")
    subprocess.run(
        [sys.executable, str(here / "annotate.py"), str(args.pdf), str(sheet),
         "-o", str(out), "--rows", str(args.build_dir / "rows.json"),
         "--xfdf", str(args.build_dir / "WORKED_EXAMPLE.xfdf")],
        check=True,
    )
    print(f"\nopen {out}")
    print("Every mapping in it is invented. It is not a submission artifact.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
